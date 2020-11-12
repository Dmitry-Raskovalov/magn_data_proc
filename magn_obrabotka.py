"""
Программа на вход принимает один или несколько файлов с измерениями магнитометра минимаг
На выходе таблица excel с данными из этих файлов.
С данными провидится минимальная обработкаЖ помечаются и удаляются повторы (оставляется последнее измерение)
        и помечается сбой нумерации измерений, ненужные измерения удаляются вручную
"""


class Source_file():
    FORMATS = {'Поле1': 'old_measures',
               'Поле2': 'old_measures',
               'Сектор': 'new_measures',
               'Автоматический': 'old_variations',
               'МВС': 'new_variations'}
   # ключевые слова по которым определяется формат файла

    def __init__(self, filename):
        self.filename = filename
        self.file = open(self.filename)
        self.format = None
        self.date = None

    def __str__(self):
        rep = 'открыт файл:' + self.filename
        return rep

    def get_format(self, line):
        """определяет формат файла измерения (новый или старый, вариации или рядовые измерения)"""
        for key in Source_file.FORMATS:
            if key in line:
                return Source_file.FORMATS[key]
        print('Ошибка.Неизвестный заголовок. Не могу определить формат файла')
        exit()

    def is_measure(line):
        """ определяет является ли текущая строка записью измерения"""
        return line[0].isdigit()

    def read_header(self, line):
        """ Читает заголовок начинающийся со строки 'Дата:'
        Парсит из него дату и формат файла
        и соотвественно изменяет свойства объекта .date и .format """
        self.date = line[-9:-1]
        self.file.readline()
        line = self.file.readline()
        if line.isspace():
            self.format = self.get_format(self.file.readline())
        else:
            self.format = self.get_format(line)

    def read_file(self, magn_db):
        """управляет чтением исходного файла
        определяет содержимое строки и передает команды на парсинг соотвествующим функциям.
        после парсинга, запускает обработку если в файле были измерения (если были только вариации обработка не запускается) """
        processing = False
        for line in self.file:
            if Source_file.is_measure(line):
                magn_db.get_data(line)
            elif 'Дата' in line:
                self.read_header(line)
                if 'measures' in self.format: processing = True  # флаг для случая когда в файле есть и измерения и вариации и файл заканчивается блоком вариаций
        if processing:
            magn_db.add_pr_pk()
            magn_db.delete_doubles()


class Database():
    DATA_FORMAT = {
        'old_measures': [
            "{'T':float(sep_line[0]), 'time':sep_line[2][:8], 'date':magn_file.date, 'name':sep_line[4][:6]}",
            '\t'],
        'new_measures': [
            "{'T':float(sep_line[0]), 'time':sep_line[1][:8], 'date':magn_file.date, 'pr':int(sep_line[3]), 'name':int(sep_line[4][1:7])}",
            '  '],
        'old_variations': [
            "{'T':float(sep_line[0]), 'time':sep_line[2][:8], 'date':magn_file.date}",
            '\t'],
        'new_variations': [
            "{'T':float(sep_line[0]), 'time':sep_line[1][:8], 'date':magn_file.date}",
            '  ']}

    # DATAFORMAT Форматы распаковки строки с данными из исходного файла в self.db.
    # Второе значение в списке - формат разделителя столбцов данных в исходном файле (табуляция или 2 пробела)

    def __init__(self, magn_file):
        self.db = {'measures': [], 'variations': [], 'wo_doubles': []}
        self.magn_file = magn_file

    def get_data(self, line):
        """извлекает данные из строки и записывает их в список"""
        sep_line = line.split(Database.DATA_FORMAT[self.magn_file.format][1])
        # Добавляет данные либо в "measures" либо в "variations" формат записи извлекает из DATAFORMAT
        self.db[self.magn_file.format[4:]].append(
            eval(Database.DATA_FORMAT[self.magn_file.format][0]))

    def add_pr_pk(self, point_in_pk=4):
        """добавляет в БД номер профиля и пикета рассчитывает их из номера точки"""
        # добавить возможность ввода шаблона вида RRKKKK (R=Pr K=Pk)
        if 'old' in self.magn_file.format:
            for rec in self.db['measures']:
                rec['pr'] = int(rec['name'][:2])
                rec['pk'] = float(rec['name'][2:]) / point_in_pk
        elif 'new' in self.magn_file.format:
            for rec in self.db['measures']:
                rec['pk'] = float(rec['name']) / point_in_pk
        else:
            print('некорректное значение формата')
            exit()

    def delete_doubles(self):
        """возвращает новую БД, без повторных измерений, оставляет из них последнее
        в основной БД, помечает удаленные измерения ключом 'doubles'
        при сбое в нумерации пикетов помечает измерение ключом 'num_error'
        Новую БД записывает в список 'wo_doubles' """

        # при нерегулярном шаге 'num_error' будет работать некорректно. подумать как реализовать
        prev_rec = self.db['measures'][0]
        j = 1
        for i in range(1, len(self.db['measures'])):
            if not (self.db['measures'][i]['pr'] == prev_rec['pr'] and
                    self.db['measures'][i]['pk'] == prev_rec['pk']):
                self.db['wo_doubles'].append(prev_rec)
                if j > 1 and (self.db['wo_doubles'][-2]['pk'] -
                              self.db['wo_doubles'][-1]['pk']) != (
                        self.db['wo_doubles'][-1]['pk'] -
                        self.db['measures'][i]['pk']):
                    if self.db['measures'][i]['pr'] == \
                            self.db['wo_doubles'][-1]['pr']:
                        self.db['measures'][i]['num_error'] = True
                    j = 0
                j += 1
            else:
                self.db['measures'][i - 1]['doubles'] = True
                if 'num_error' in self.db['measures'][
                    i - 1]:  # условие для случая когда на пикете произошел сбой и повтор одновременно, что бы метка осталась в финальной БД
                    del self.db['measures'][i - 1][
                        'num_error']  # пометку о сбое нумерации удаляем что бы каждой записи соответствовал один ключ при экспорте таблицы в excel не возникало конфликта при раскраске
                    self.db['measures'][i]['num_error'] = True
            prev_rec = self.db['measures'][i]
        self.db['wo_doubles'].append(prev_rec)


class Excel_book():
    SHEET_NAMES = {'wo_doubles': 'с_пикетами', 'measures': 'исходник',
                   'variations': 'вариации'}
    DATA_COLUMN_FORMAT = {
        'measures': [['T', 'Время', 'Дата', 'ПР', 'ПК_прибор', 'примечания'],
                     ['T', 'time', 'date', 'pr', 'name']],
        'variations': [['T', 'Время', 'Дата'], ['T', 'time', 'date']],
        'wo_doubles': [
            ['T', 'Время', 'Дата', '№т./пк_прибор', 'ПР', 'ПК', 'примечания'],
            ['T', 'time', 'date', 'name', 'pr', 'pk']]}

    def __init__(self, filename, is_add):
        self.filename = filename[0]
        self.wb = self.load_excel_book() if is_add else self.create_excel_book()

    def create_excel_book(self):
        """создает файл excel c несколькими листами количество и назание берется из SHEET_NAMES """
        import openpyxl
        wb = openpyxl.Workbook()
        for key in Excel_book.SHEET_NAMES:
            wb.create_sheet(title=Excel_book.SHEET_NAMES[key])
        sheet_list = wb.sheetnames
        print(sheet_list)
        wb[sheet_list[0]].cell(row=1, column=1).value = 1
        del wb[sheet_list[0]]
        wb.save(self.filename)
        print('файл ', self.filename, 'создан')
        return wb

    def load_excel_book(self):
        import openpyxl
        wb = openpyxl.load_workbook(self.filename)
        sheet_list = wb.sheetnames
        for sheet_name in Excel_book.SHEET_NAMES.values():
            if not sheet_name in sheet_list:
                print('в выбранном файле нет листа :', sheet_name)
                exit()
        else:
            print('файл', self.filename, 'успешно открыт')
        return wb

    def write_db_to_excel_book(self, sources_db):
        """ Управляет записью данных в таблицу. Непосредственно запись производится функцией write_sheet.
        Здесь определяеся содержание и последовательность столбцов в таблице.
        Формат таблиц одинаков для старых и новых магнитометров. Дисперсия (для старых магнитометров) опускается.
         """
        import openpyxl
        from openpyxl.styles import PatternFill
        for key in Excel_book.SHEET_NAMES:
            self.write_sheet(key, sources_db.db[key])
        self.wb.save(self.filename)
        print('Данные добавлены в файл: ', self.filename)

    def write_sheet(self, key, db):
        """ Записывает данные, прочитанные из исходного файла в таблицу.
        Если таблица пустая добавляет заголовок.
        Раскрашивает ячейки с ошибками"""
        import openpyxl
        from openpyxl.styles import PatternFill
        FLAGS = [{'mark': 'num_error', 'sence': 'Сбой нумерации',
                  'cells_color': 'ff9966'},
                 {'mark': 'doubles', 'sence': 'Повтор'	, 'cells_color' :'cc9999'}]
        work_sheet = self.wb[Excel_book.SHEET_NAMES[key]]
        data_header = Excel_book.DATA_COLUMN_FORMAT[key][0]
        data_format = Excel_book.DATA_COLUMN_FORMAT[key][1]
        if work_sheet.max_row < 2:
            work_sheet.append(data_header)
            i = 2
        else:
            i = work_sheet.max_row + 1
        for row in db:
            row_to_write = []
            for each in data_format:
                row_to_write.append(row[each])
            work_sheet.append(row_to_write)
            for flag in FLAGS:
                if flag['mark'] in row:
                    work_sheet.cell(row=i, column=work_sheet.max_column).value = flag['sence']
                    for col_num in range(1, work_sheet.max_column + 1):
                        work_sheet.cell(row=i, column=col_num).fill = PatternFill \
                            (fill_type='solid', start_color=flag['cells_color'])
            i += 1


class File_choise():
    """текстовое меню для выбора файлов"""

    def __init__(self, file_extention=''):
        self.file_list = self.take_file_list(file_extention)
        self.is_add = False
        self.selected_files = []  # для реализации выбора нескольких файлов
        self.file_extention = file_extention

    def get_input_txt(self):
        """ выводит список txt файлов и принимает пользовательский выбор одного или нескольких из них
        пользователь вводит номер файла в списке"""
        import re
        file_name = None
        menu_str = re.compile(r'^(\s)*?\d+((\s)*?,(\s)*?\d+)*$')
        if self.file_list:
            self.print_list()
            while not file_name:
                print \
                    ('Для выбора исходного файла введите его номер в списке.\n Можно перечислить несколько файлов через запятую.')
                print('Для выхода введите q')
                menu_item = input('\n >')
                if menu_item.lower() == 'q': exit()
                if not menu_str.search \
                    (menu_item) or '@' in menu_item: continue
                file_name = self.check_input(menu_item)
                print('выбран(ы) файл(ы): ')
                for file in file_name: print(file)
                print('- ' *20)
        else:  # когда в директории нет файлов c нужным расширением
            print('в этом каталоге нет подходящих файлов')
            exit()
        return file_name

    def get_input_xlsx(self):
        """Выводит список xlsx файлов. Предлагает пользователю выбрать один из них для добавления к нему данных,
        либо ввести имя нового файла.
        """
        import re
        menu_str = re.compile(r'(^(\s)*?\d+(\s)*?)|(@..*)')
        # добавить проверку на корректность имени файла с помощью регулярных выражений
        file_name = None
        self.print_list()
        while not file_name:
            print \
                ('если хотите добавить данные в существующий файл введите номер файла. \nЛибо введите имя нового файла после знака @'
                  if self.file_list else 'В этом каталоге нет файлов .xlsx. Введите имя нового файла после знака @.')
            print('Для выхода введите q')
            menu_item = input('\n >')
            if menu_item.rstrip().lower() == 'q': exit()
            if not menu_str.search(menu_item): continue
            file_name = self.check_input(menu_item)
        return file_name

    def take_file_list(self, file_extention=''):
        """Принимает на вход строку вида '.txt' c требуемым раcширением файлов
         и возвращает список файлов в текщей папке с таким расширением """
        import os
        return  [filename for filename in os.listdir(r'd:\my_code\magnitka_obrabotka\\') if file_extention in filename]

    def print_list(self, per_screen=25):
        """ выводит построчно список файлов с номерами по per_screen шт. с паузами"""
        for i, filename in enumerate(self.file_list):
            print(i, ' ', filename)
            if i != 0 and i % per_screen == 0 and i + 1 < len \
                    (self.file_list):  # Длинный список файлов будет выводится по per_screen файлов с паузами
                input('нажмите enter для продолжения')

    def check_input (self, choice, is_list=False):
        """ Проверяет корректность пользовательского ввода (пункта меню) или имени файла.
        Если выбран пункт меню передает флаг(is_add) на добавление  данных в файл, если введено имя файла - флаг записи в новый.
        если введено имя существующего файла - флаг на добавление данных в него """
        file_list = self.file_list
        menu_items = []
        file_name = []
        if not '@' in choice[0]:
            self.is_add = True
            if ',' in choice:
                menu_items = eval(choice)
            else:
                menu_items.append(int(choice))
            for item in menu_items:
                if item in range(0, len(file_list)):
                    file_name.append(file_list[item])
                else:
                    print('число должно быть от 0 до ', len(file_list) - 1)
        else:
            self.is_add = False
            if '.xlsx' in choice[-5:]:
                file_name.append(choice[1:])
            else:
                file_name.append(choice[1:] + '.xlsx')
            if file_name[0] in file_list:
                self.is_add = True
        return file_name

def exit():
    import sys
    input('нажмите enter для выхода')
    sys.exit()


txt_file = File_choise('.txt')
xlsx_file = File_choise('.xlsx')
files_name = txt_file.get_input_txt()
magn_book = Excel_book(xlsx_file.get_input_xlsx(), xlsx_file.is_add)
for file in files_name:
    magn_file = Source_file(file)
    magn_db = Database(magn_file)
    magn_file.read_file(magn_db)
    magn_book.write_db_to_excel_book(magn_db)

"""
print(magn_file)
print('дата:', magn_file.date)
print('формат: ', magn_file.format)
for key in magn_db.db:
	print(key)
	for rec in magn_db.db[key]: print(rec)
"""

# Баг: если в файле есть разрыв (магнитометр отключался на профиле) и в этом разрыве есть лишний заголовок и с однимп значением, то программа записывает это значение в вариации
# добавить сохранение бэкапа
# добавить другие варианты нумерации пикетов (ввод вручную начального-конечного),
# добавления таблицы соответствия пикетов, пометку о наличии регулярного шага между пикетами
# если шаг между пикетами равномерный - проверка кооректности ввода номеров пикетов




